import warnings 
import openpyxl
from pymongo import MongoClient

warnings.simplefilter("ignore")
wb = openpyxl.load_workbook("order_data.xlsx")
sheet = wb.get_sheet_by_name("Sheet1")

client = MongoClient('localhost:27017')
db = client.shadowfax
for rowNum in range(2, sheet.get_highest_row()):
    order={}
    order["order_id"]=sheet.cell(row=rowNum, column=1).value
    order["seller_id"]=sheet.cell(row=rowNum, column=2).value
    order["rider_id"]=sheet.cell(row=rowNum, column=3).value
    order["cluster_id"]=sheet.cell(row=rowNum, column=4).value
    if sheet.cell(row=rowNum, column=5).value is not None:
        sheet.cell(row=rowNum, column=5).value=sheet.cell(row=rowNum, column=5).value.replace("\"","")
    order["scheduled_time"]=sheet.cell(row=rowNum, column=5).value

    if sheet.cell(row=rowNum, column=6).value is not None:
        sheet.cell(row=rowNum, column=6).value=sheet.cell(row=rowNum, column=6).value.replace("\"","")
    order["allot_time"]=sheet.cell(row=rowNum, column=6).value

    if sheet.cell(row=rowNum, column=7).value is not None:
        sheet.cell(row=rowNum, column=7).value= order["pickup_time"]=sheet.cell(row=rowNum, column=7).value.replace("\"","")
    order["pickup_time"]=sheet.cell(row=rowNum, column=7).value

    if sheet.cell(row=rowNum, column=8).value is not None:
        sheet.cell(row=rowNum, column=8).value=sheet.cell(row=rowNum, column=8).value.replace("\"","")
    order["delivered_time"]=sheet.cell(row=rowNum, column=8).value
    
    order["pickup_latitude"]=sheet.cell(row=rowNum, column=9).value
    order["pickup_longitude"]=sheet.cell(row=rowNum, column=10).value
    order["delivered_latitude"]=sheet.cell(row=rowNum, column=11).value
    order["delivered_longitude"]=sheet.cell(row=rowNum, column=12).value
    order["house_number"]=sheet.cell(row=rowNum, column=13).value
    order["sublocality"]=sheet.cell(row=rowNum, column=14).value
    db.order_data.insert(order)
print (rowNum)
print("Done")

